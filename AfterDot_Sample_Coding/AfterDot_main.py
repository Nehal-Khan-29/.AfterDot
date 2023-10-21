'''
project done by 
NEHAL KHAN
'''

import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog as fd
from tkinter.filedialog import askdirectory
from tkinter import ttk as ttk
from datetime import date
import time
import os
from PIL import ImageTk,Image
from PIL import Image
import pandas as pd
import openpyxl
import pytesseract
#import subprocess
import ctypes
#import sys
import pdf2docx
from pdf2docx import Converter
from pdf2docx import parse
from typing import Tuple
from docx2pdf import convert
import docx
from docx import Document
from PyPDF2 import PdfMerger, PdfReader
from pdf2image import convert_from_path
import PyPDF2



# Page Close Confirmations (Messagebox):

def homelogout():
    messagebox.showinfo('Thank You','Logged out successfully')
    home.destroy()
        
def homeclose():
    if messagebox.askokcancel('Quit','Do you want to logout and quit?'):
        home.destroy()
        quit()





# icon window

icon = tk.Tk()
icon.title('.AfterDot                                THE BEST CONVERTER')
icon.iconbitmap("E:\\NK programs\\Python\\python save\\AfterDot\\icon ICO.ico")
image = Image.open("E:\\NK programs\\Python\\python save\\AfterDot\\icon.png")
tk_image = ImageTk.PhotoImage(image)
image_label = tk.Label(icon, image=tk_image)
image_label.pack()
icon.update()
screen_width = icon.winfo_screenwidth()
screen_height = icon.winfo_screenheight()
window_width = 520  
window_height = 450  
x = int((screen_width - window_width) / 2)
y = int((screen_height - window_height) / 2)
icon.geometry("+{}+{}".format(x, y))
icon.after(2000, icon.destroy)
icon.mainloop()



#Tesseract
def tes():
    directory = "C:\\Program Files"
    filename = "Tesseract-OCR"
    filepatht = os.path.join(directory, filename)

    if not os.path.exists(filepatht):
        ctypes.windll.shell32.ShellExecuteW(None, "runas", "E:\\NK programs\\Python\\python save\\AfterDot\\tesseract-ocr-w64-setup-5.3.1.20230401.exe", None, None, 1)
        print(f"Tesseract file '{filename}' will be created in '{directory}' directory after the installation.")
    else:
        print(f"Tesseract file '{filename}' already exists in '{directory}' directory.")





#Excel add
directory = "E:\\NK programs\\Python\\python save\\AfterDot"
filename = "AfterDot Points.xlsx"
filepathx = os.path.join(directory, filename)

if not os.path.exists(filepathx):
    df = pd.DataFrame({'Sno': [1,2,3,4,5,6,7,8,9,10], 'Func': ['word to pdf','pdf to word','pdf to word 2','pdf to text','pdf merger','pdf to jpg','jpg to pdf','jpg to png','png to jpg','image to text'],'points': [0,0,0,0,0,0,0,0,0,0]})
    with pd.ExcelWriter(filepathx) as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    print(f"Excel file '{filename}' created in '{directory}' directory.")
else:
    print(f"Excel file '{filename}' already exists in '{directory}' directory.")





# Word To PDF Page:

def wordtopdf(): 
    filetype=('docx files','*.docx'),('all files','*.*')
    filepath=fd.askopenfilenames(filetypes=filetype)
    a=filepath

    z=a[0]
    b=z[-6:-len(z)-1:-1]
    c=b[-1:-len(b)-1:-1]

    if z[-5:] == '.docx':
        convert(f"{c}.docx",
        f"{c}.pdf")

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 2
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)



   


# PDF To Word Page:

def pdftoword():
    filetype=('pdf files','*.pdf'),('all files','*.*')
    filepath=fd.askopenfilenames(filetypes=filetype)
    a=filepath

    z=a[0]
    b=z[-5:-len(z)-1:-1]
    c=b[-1:-len(b)-1:-1]

    #parse(f"{c}.pdf",f"{c}.docx")
    pdf2docx.parse(f"{c}.pdf", f"{c}.docx", headers=True, footers=True)

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 3
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)






# Pdf to word 2:

def pdftoword2():
    pdf_filepath = fd.askopenfilename(filetypes=(('PDF files', '*.pdf'), ('All files', '*.*')))

    with open(pdf_filepath, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        pdf_text = ''
        for page in range(len(pdf_reader.pages)):
            pdf_text += pdf_reader.pages[page].extract_text()

    doc = docx.Document()
    doc.add_paragraph(pdf_text)
    docx_filepath = pdf_filepath.replace('.pdf', '.docx')
    doc.save(docx_filepath)

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 4
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)

  




#pdf to text:

def pdftotxt():
    root = tk.Tk()
    root.withdraw()

    filetypes = [('PDF files', '*.pdf'), ('All files', '*.*')]
    filepaths = fd.askopenfilenames(filetypes=filetypes)

    if len(filepaths) > 0:
        filepath = filepaths[0]

        with open(filepath, 'rb') as pdf_file:

            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
            text = ""

            for i in range(num_pages):
                page = pdf_reader.pages[i]
                text += page.extract_text()

       

        filename_without_ext = os.path.splitext(os.path.basename(filepath))[0]

        txt_filepath = os.path.join(os.path.dirname(filepath), f"{filename_without_ext}.txt")
        with open(txt_filepath, 'w', encoding='utf-8') as txt_file:
            txt_file.write(text)

        '''
        with open(txt_filepath, 'rb') as file:          
            file_contents = file.read()
        new_contents = file_contents.decode('utf-8').replace('apple', 'Nehal')
        with open(txt_filepath, 'w') as file:
            file.write(new_contents)'''

        

    else:
        print("No files selected")

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 5
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)






# PDF To JPG Page:

def pdftojpg():

    filetype=('pdf files','*.pdf'),('all files','*.*')
    filepath=fd.askopenfilenames(filetypes=filetype)
    a=filepath

    z=a[0]
    b=z[-5:-len(z)-1:-1]
    c=b[-1:-len(b)-1:-1]

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 7
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)

    
    images = convert_from_path(f"{z}",500,poppler_path=r'E:\NK programs\Python\python save\AfterDot\poppler-23.01.0\Library\bin')
    
    for i in range(len(images)):
	    images[i].save(f"{c}"+' page'+ str(i+1) +'.jpg', 'JPEG')
            
    





# JPG To PDF Page:

def jpgtopdf():

    filetype=('jpg files','*.jpg'),('all files','*.*')
    filepath=fd.askopenfilenames(filetypes=filetype)
    a=filepath

    z=a[0]
    b=z[-5:-len(z)-1:-1]
    c=b[-1:-len(b)-1:-1]
    
    image = Image.open(z,'r')
    im1 = image.convert('RGB')
    im1.save(f"{c}.pdf")

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 8
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)






#multiple jpd to pdf
def multijpgtopdf():
    print()






# 2xPDF TO PDF

def pdftopdf():
    filetype1=('pdf files','*.pdf'),('all files','*.*')
    filepath1=fd.askopenfilenames(filetypes=filetype1)
    a=filepath1
    filetype2=('pdf files','*.pdf'),('all files','*.*')
    filepath2=fd.askopenfilenames(filetypes=filetype2)
    x=filepath2

    z=a[0]
    y=x[0]
    b=z[-5:-len(z)-1:-1]
    c=b[-1:-len(b)-1:-1]

    merger=PdfMerger()
    merger.append(z)
    merger.append(y)
    merger.write(c+" merged.pdf")
    merger.close()

    wb = openpyxl.load_workbook(filepathx)
    sheet = wb.active
    row_num = 6
    col_num = 3
    old_value = sheet.cell(row=row_num, column=col_num).value
    new_value = old_value +1
    sheet.cell(row=row_num, column=col_num).value = new_value
    wb.save(filepathx)


#multiple pdf merger
def multipdf():
    print()


# PNG TO JPG

def pngtojpg():
    print()




# multi PNG TO JPG

def multipngtojpg():
    print()








# JPG TO PNG

def jpgtopng():
    print()




#multi jpg to png

def multijpgtopng():
    print()


#image to txt

def imgtotxt():
    print()





# Points Page:

def pointspage():

    point=tk.Toplevel()
    point.geometry('1366x768')
    point.title('AfterDot - Points')
    point.iconbitmap("E:\\NK programs\\Python\\python save\\AfterDot\\icon ICO.ico")
    point.state('zoomed')

    pointpic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\AfterDot\\afterdot front 2.png"))
    pointpanel=Label(point,image=pointpic)
    pointpanel.pack(side='top',fill='both',expand='yes')

    df = pd.read_excel("E:\\NK programs\\Python\\python save\\AfterDot\\AfterDot Points.xlsx")

    tree=ttk.Treeview(point,column=('#c1','#c2','#c3'),show='headings',height=10)

    tree.column('#1',width=140,minwidth=140,anchor=tk.CENTER)
    tree.column('#2',width=140,minwidth=140,anchor=tk.CENTER)
    tree.column('#3',width=140,minwidth=140,anchor=tk.CENTER)

    tree.heading('#1',text='sno')
    tree.heading('#2',text='func')
    tree.heading('#3',text='points')
    tree.pack()
    
    for index, row in df.iterrows():
        tree.insert('', 'end', values=tuple(row))
    tree.place(x=580,y=220)
    

    def reset():
        for i in range(2,12):
            wb = openpyxl.load_workbook(filepathx)
            sheet = wb.active
            row_num = i
            col_num = 3
            new_value = 0
            sheet.cell(row=row_num, column=col_num).value = new_value
            wb.save(filepathx)
            point.destroy()   

    Button(point,text='Reset points',font=('Arial',20),command=reset,height=1,width=18,bg='red',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=1030,y=580)
    
    point.mainloop()





# instruction page:

def instructionpage():

    instruc=tk.Toplevel()
    instruc.geometry('1366x768')
    instruc.iconbitmap("E:\\NK programs\\Python\\python save\\AfterDot\\icon ICO.ico")
    instruc.title('AfterDot - About')
    instruc.state('zoomed')

    instrucpic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\AfterDot\\afterdot front 2.png"))
    instrucpanel=Label(instruc,image=instrucpic)
    instrucpanel.pack(side='top',fill='both',expand='yes')


    Label(instruc,text=('''>>> Single file conversion:

Choose the function from the given choices, after that a File dialog box appears, asking you to choose your desired file to convert.
Once the file gets converted, the file appears with the same name in the same directory as the original file was located.

For the PDF Merger, the same steps follow, but 2 File dialog box appears for each 2 PDFs which is about to be merged.

>>> Multi-file conversion:

Here only 1 File dialog box appears, and you have to choose all the desired files in 1 go.

>>> Methods to choose files in 1 go:

1. Press and hold CTRL and left click with mouse each file in order (use this if you are going to skip any file to convert)
2. Press and hold CTRL and SHIFT and LEFT/RIGHT/UP/DOWN arrow (use this if you are not going to skip any file to convert)
3. Select files by RIGHT CLICK and HOLD ()
4. CTRL + A to select all files in folder ()

>>> Important Note:

use 2,3,4 only if your desired file order is matching the actual order in the folder where the files 
are located or else the pages in your PDF MERGER will be unordered'''),font=('Arial',15),bg='white', justify='left').place(x=120,y=170)
    
    instruc.mainloop()


# About Page:

def aboutpage():

    about=tk.Toplevel()
    about.geometry('1366x768')
    about.iconbitmap("E:\\NK programs\\Python\\python save\\AfterDot\\icon ICO.ico")
    about.title('AfterDot - About')
    about.state('zoomed')

    aboutpic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\AfterDot\\afterdot front 2.png"))
    aboutpanel=Label(about,image=aboutpic)
    aboutpanel.pack(side='top',fill='both',expand='yes')

    def help():
        Label(about,text='''MAIL ID                         : nehalmicro29@gmail.com''',font=('Arial',16)).place(x=750,y=550)

    Button(about,text='Contact Us',font=('Arial',20),command=help,height=1,width=16,bg='white',
    fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=750,y=500)


    Label(about,text=(''' Convert them as many times as you like !!! '''),
                 font=('Arial',12),bg='white',borderwidth=0,relief='solid').place(relx=0.41,y=630)

    about.mainloop()





# Home Page:

home=tk.Tk()
home.geometry('1366x768')
home.iconbitmap("E:\\NK programs\\Python\\python save\\AfterDot\\icon ICO.ico")
home.title('AfterDot')
home.state('zoomed')
home.protocol('WM_DELETE_WINDOW',homeclose)

currtime=time.strftime('%H:%M')
currdate=date.today().strftime("%d/%m/%Y")

homepic=ImageTk.PhotoImage(Image.open("E:\\NK programs\\Python\\python save\\AfterDot\\afterdot front 2.png"))
homepanel=Label(home,image=homepic)
homepanel.pack(side='top',fill='both',expand='yes')

Label(home,text=('Logged in: '+currtime+' - '+currdate),font=('Arial',16),bg='blue',fg="white").place(x=1165,y=80)




'''BLINK_COLOR = "blue"
def blink_on_enter(event):
    event.widget.config(bg=BLINK_COLOR)
Button.bind("<Enter>", blink_on_enter)'''



# horizontal 1

Label(home,text=("Single select file conversion "),font=('Arial',16),bg='white',fg="blue").place(x=637,y=125)

Button(home,text='Word -> PDF',font=('Arial',20),command=wordtopdf,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=100,y=200)
Button(home,text='PDF -> Word',font=('Arial',20),command=pdftoword,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=200)
Label(home,text=("Only borderless pdf"),font=('Arial',16),bg='white',fg="blue").place(x=500,y=255)
Button(home,text='PDF -> Word 2',font=('Arial',20),command=pdftoword2,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=820,y=200)
Label(home,text=("Only bordered pdf"),font=('Arial',16),bg='white',fg="blue").place(x=870,y=255)
Button(home,text='PDF -> TEXT',font=('Arial',20),command=pdftotxt,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=1180,y=200)


# horizontal 2

Button(home,text='PDF merger',font=('Arial',20),command=pdftopdf,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=100,y=300)
Button(home,text='PDF -> JPG',font=('Arial',20),command=pdftojpg,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=300)
Button(home,text='JPG -> PDF',font=('Arial',20),command=jpgtopdf,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=820,y=300)
Button(home,text='JPG -> PNG',font=('Arial',20),command=jpgtopng,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=1180,y=300)

# horizontal 3

Button(home,text='PNG -> JPG',font=('Arial',20),command=pngtojpg,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=100,y=400)
Button(home,text='IMG -> TEXT',font=('Arial',20),command=imgtotxt,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=400)


Label(home,text=("Multi select file conversion "),font=('Arial',16),bg='white',fg="blue").place(x=650,y=478)


# horizontal 4

Button(home,text='PDF merger',font=('Arial',20),command=multipdf,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=100,y=530)
Button(home,text='JPG -> PDF',font=('Arial',20),command=multijpgtopdf,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=530)
Button(home,text='JPG -> PNG',font=('Arial',20),command=multijpgtopng,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=820,y=530)
Button(home,text='PNG -> JPG',font=('Arial',20),command=multipngtojpg,height=1,width=16,bg='white',
       fg='gray6',activebackground='Skyblue',activeforeground='thistle1').place(x=1180,y=530)

# horizontal 5

Button(home,text='Points',font=('Arial',20),command=pointspage,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=100,y=630)
Button(home,text='Instruction',font=('Arial',20,),command=instructionpage,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=460,y=630)
Button(home,text='About Us',font=('Arial',20,),command=aboutpage,height=1,width=16,bg='blue',
       fg='white',activebackground='Skyblue',activeforeground='thistle1').place(x=820,y=630)
Button(home,text='Logout',font=('Arial',20),command=homelogout,height=1,width=16,bg='lightgreen',
       fg='red',activebackground='Skyblue',activeforeground='thistle1').place(x=1180,y=630)



home.mainloop()


